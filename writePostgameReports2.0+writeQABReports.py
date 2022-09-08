import csv
import os
import sqlite3
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import pandas as pd
import keyboard
import time
import pyautogui as mouse


path = "c:\\Users\\rlars\Desktop\Knights Baseball\\2022 WCL Trackman CSVs\\August\\August 15\\20220815-GossStadium-1_unverified.csv"
lineup = ['Romero, Kiko', 'Scott, Spencer', 'Advincula, Jonah', 'Kennel, Ely', 'Johnstone, Logan', 'Loveless, Ethan', 'Dumitru, Titus', 'Casperson, Kyle',
'Becerra, Temo', 'Darby, Zander', 'Le, Mason', 'Lavoie, Brady', 'Knight, Briley', 'DiPaolo, Luca', 'Quinn, Tyler', 'Stafford, Ryan', 'Smith, Tanner']

pitchers = ['Scott, Matt', 'Segel, Kaden', 'Ross, Ethan', 'Day, Cam']


autoTagPitchType = False #True means to take the autoTag type becuase the other WCL Trackman taggers are shitty

# Put Cowlitz on the list of shitty Trackman admins
adjustmentY = 0 #Bottom of the zone is around 1.65, Walla Walla required about a +1.5 adjustment
adjustmentX = 0


dbName, gamedate, opponent, tableName = 'Knights2022', '2022-08-15', 'Bellingham', 'August15vBellingham'

def pitchType(autoTagPitchType):
    out = 'taggedPitch'
    if autoTagPitchType:
        out = 'autoPitch'
    return out

def twoNameFix(player):
    #Trackman name converts to Pointstreak name
    if player == 'Segal, Kaden':
        player = 'Segel, Kaden'
    elif player == 'Vargas, Jacob':
        player = 'Vargas, Jake'
    elif player == 'Mendoza, Nate':
        player = 'Mendoza, Nathaniel'
    elif player == 'Deschryver, Nathan':
        player = 'DeSchryver, Nathan'
    elif player == 'Walter, Chase Ryan':
        player = 'Walter, Chase'
    elif player == 'Nathan Nankil':
        player = 'Nate Nankil'
    elif player == 'Jacob Hoskins':
        player = 'Jake Hoskins'
    elif player == 'Kaden Segal':
        player = 'Kaden Segel'
    elif player == 'Meyer, Colton ':
        player = 'Meyer, Colton'
    elif player == 'Garcia Jr, Sebastian':
        player = 'Garcia, Sebastian'
    elif player == 'Ethan Lovelass':
        player = 'Ethan Loveless'
    elif player == 'Dobmeier, Jacob ':
        player = 'Dobmeier, Jacob'
    elif player =='Ager, Matt':
        player = 'Ager, Matthew'
    return player

def getRows(filename):
    with open(filename, 'r') as csvfile:
        csvreader = csv.reader(csvfile)
        rows = {}
        i = 0
        for row in csvreader:
             i += 1
             rows[i] = row
    return rows

def getAtBats(path):
    rows = getRows(path)
    heading = rows[1]

    hitter = heading.index('Batter')
    currentHitter = rows[2][hitter]
    atBats = {}
    a = 1
    for i in range(2, len(rows)+1):
        if currentHitter == rows[i][hitter]:
            if a in atBats:
                atBats[a].append(rows[i])
            else:
                atBats[a] = [rows[i]]
        else:
            a += 1
            atBats[a] = [rows[i]]
            currentHitter = rows[i][hitter]
    return atBats

def getIndex(path):
    rows = getRows(path)
    heading = rows[1]

    out = {'date' : heading.index('Date'),
    'time' : heading.index('Time'),
    'ab' : heading.index('PAofInning'),
    'inning' : heading.index('Inning'),
    'hitter' : heading.index('Batter'),
    'topBottom' : heading.index('Top/Bottom'),
    'autoPitchType' : heading.index('AutoPitchType'),
    'taggedPitchType' : heading.index('TaggedPitchType'),
    'pitcher' : heading.index('Pitcher'),
    'outs' : heading.index('Outs'),
    'pitchResult' : heading.index('PitchCall'),
    'exitVelo' : heading.index('ExitSpeed'),
    'launchAngle' : heading.index('Angle'),
    'distanceTraveled' : heading.index('Distance'),
    'hitType' : heading.index('TaggedHitType'),
    'autoHitType' : heading.index('AutoHitType'),
    'playOutcome' : heading.index('PlayResult'),
    'pitchLocationX' : heading.index('PlateLocSide'),
    'pitchLocationY' : heading.index('PlateLocHeight'),
    'spinRate' : heading.index('SpinRate'),
    'pitchVelo' : heading.index('RelSpeed'),
    'inducedBreak' : heading.index('InducedVertBreak'),
    'vertBreak' : heading.index('VertBreak'),
    'horzBreak' : heading.index('HorzBreak'),
    'strikes' : heading.index('Strikes'), 
    'balls' : heading.index('Balls'), 
    'batterSide' : heading.index('BatterSide'),
    'paInning' : heading.index('PAofInning'),
    'pitchPA' : heading.index('PitchofPA'),
    'pitcherSide' : heading.index('PitcherThrows'),
    'KorBB' : heading.index('KorBB'),
    'runsScored' : heading.index('RunsScored')}

    return out
    
def openDatabase(name):
    path = "c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\Databases\Knights2022"
    con = sqlite3.connect(path)
    cur = con.cursor()

    return cur, con

def writeDatabase(name, tableName, path, aX = adjustmentX, aY = adjustmentY):
    cur, con = openDatabase(name)
    atBats = getAtBats(path)
    indexDic = getIndex(path)

    cur.execute(f'''CREATE TABLE IF NOT EXISTS {tableName} (pitchNum INTEGER PRIMARY KEY, date TEXT, time TEXT, ab INTEGER, paInning INTEGER, pitchPA INTEGER, inning INTEGER, topBottom TEXT, hitter TEXT, batterSide TEXT,
    autoPitch TEXT, taggedPitch TEXT, pitcher TEXT, pitcherSide TEXT, outs NUMBER, pitchResult TEXT, exitVelo NUMBER, launchAngle NUMBER, distance NUMBER,
    hitType TEXT, autoHitType TEXT, outcome TEXT, KorBB TEXT, runsScored INTEGER, locationX NUMBER, locationY NUMBER, spinRate NUMBER, pitchVelo NUMBER, inducedBreak NUMBER, vertBreak NUMBER, horzBreak NUMBER, balls INTEGER, strikes INTEGER)''')
    con.commit()

    pitchNum = 1
    for atBat in atBats:
        for pitch in atBats[atBat]:
            try:
                pitchX = float(pitch[indexDic['pitchLocationX']]) + aX
                pitchY = float(pitch[indexDic['pitchLocationY']]) + aY
            except ValueError:
                print(f'No spin and/or velo recorded: AB {pitch[indexDic["ab"]]} of inning {pitch[indexDic["inning"]]}')
                pitchX = ''
                pitchY = ''

            tup =(pitchNum, pitch[indexDic['date']], pitch[indexDic['time']], pitch[indexDic['ab']], pitch[indexDic['inning']], pitch[indexDic['topBottom']], pitch[indexDic['hitter']],
            pitch[indexDic['autoPitchType']], pitch[indexDic['taggedPitchType']], pitch[indexDic['pitcher']], pitch[indexDic['outs']], pitch[indexDic['pitchResult']],
            pitch[indexDic['exitVelo']], pitch[indexDic['launchAngle']], pitch[indexDic['distanceTraveled']], pitch[indexDic['hitType']], pitch[indexDic['autoHitType']],
            pitch[indexDic['playOutcome']], pitch[indexDic['KorBB']], pitch[indexDic['runsScored']], pitchX, pitchY, pitch[indexDic['spinRate']], pitch[indexDic['pitchVelo']], pitch[indexDic['inducedBreak']], pitch[indexDic['vertBreak']],
            pitch[indexDic['horzBreak']], pitch[indexDic['balls']], pitch[indexDic['strikes']], pitch[indexDic['batterSide']], pitch[indexDic['paInning']], pitch[indexDic['pitchPA']], pitch[indexDic['pitcherSide']])
            
            cur.execute(f'''INSERT INTO {tableName} (pitchNum, date, time, ab, inning, topBottom, hitter, autoPitch, taggedPitch, pitcher, outs, pitchResult,
            exitVelo, launchAngle, distance, hitType, autoHitType, outcome, KorBB, runsScored, locationX, locationY, spinRate, pitchVelo, inducedBreak, vertBreak, horzBreak, balls, strikes, batterSide, paInning, pitchPA, pitcherSide) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''', tup)
            con.commit()
            pitchNum += 1
 
def openHitterWorkbook(copy = 'Knights Hitter Postgame Template 2.0.xlsx'):
    copyPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\Templates\{copy}'
    wb = load_workbook(copyPath)
    return wb.active, wb

def openPitcherWorkbook(copy = 'Knights Pitcher Postgame Template Copy.xlsx'):
    copyPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\Templates\{copy}'
    wb = load_workbook(copyPath)
    return wb.active, wb

def writePitches(pitchVars, player, gamedate, sheet, cur, tableName):

    cur.execute(f'SELECT {pitchVars} FROM {tableName} WHERE hitter = ? AND date = ?' , (player, gamedate))
    allPitches = cur.fetchall()

    pitchTypeDic = {'Fastball' : ('FB', 'Red', '00FF0000') , 'Four-Seam': ('FB', 'Red', '00FF0000'), 'ChangeUp' : ('CH', 'Blue', '000000FF'), 'Changeup' : ('CH', 'Blue', '000000FF'),'Slider' : ('SL', 'Green', '00339966') , 'Cutter' : ('CUT', 'Turquoise', '0040E0D0'),
     'Curveball' : ('CB', 'LimeGreen', '0000FF00') , 'Splitter' : ('SP', 'LightBlue', '0000CCFF'), 'Sinker' : ('2FB', 'Orange', '00FF6600'), 'Knuckleball' : ('KN', 'DarkGreen', '00008000'), 'Undefined' : ('N/A', 'White', '00FFFFFF'), '' : ('N/A', 'White' ,'00FFFFFF'), 'Other' : ('N/A', 'White' ,'00FFFFFF')}
    pitchResultDic = {'BallCalled' : 'Ball' , 'StrikeCalled' : 'Strike' , 'FoulBall' :'Foul', 'InPlay' : 'In Play', 'StrikeSwinging' : 'Swing', 'HitByPitch' : 'HBP', 'BallinDirt' : 'Ball', 'Undefined' : 'Error'}

    abCount = 1
    count = 10
    try:
        num = allPitches[0][0]
    except IndexError:
        #print(f'Exception while pulling from database: {player}')
        pass
    for pitch in allPitches:
        if count >= 50:
            break
        if pitch[0] == num:
            count = str(count)
            if not pitch[-1]:
                sheet[f'F{count}'] = 'Error'
                sheet[f'F{count}'].font = Font(color = '00000000')
                sheet[f'G{count}'] = 'Error'
            else:
                sheet[f'F{count}'] = pitchTypeDic[pitch[1]][0]
                sheet[f'F{count}'].font = Font(color = pitchTypeDic[pitch[1]][2])
                sheet[f'G{count}'] = pitchResultDic[pitch[2]]
            count = int(count) + 1
            num += 1
        else:
            abCount += 1
            count =  8 * abCount + 2
            if count >= 50:
                count += 7
            num = pitch[0]
            if not pitch[-1]:
                sheet[f'F{count}'] = 'Error'
                sheet[f'F{count}'].font = Font(color = '00000000')
                sheet[f'G{count}'] = 'Error'
            else:
                sheet[f'F{count}'] = pitchTypeDic[pitch[1]][0]
                sheet[f'F{count}'].font = Font(color = pitchTypeDic[pitch[1]][2])
                try:
                    sheet[f'G{count}'] = pitchResultDic[pitch[2]]
                except:
                    sheet[f'G{count}'] = 'Error'
            count = int(count) + 1
            num += 1

def drawPitchBreakChart(date, pitcher, autoTagPitchType, dbName, tableName):
    cur, con = openDatabase(dbName)

    months = {1 : 'Jan', 2: 'Feb' , 3: 'Mar' , 4: 'Apr' , 5 : 'May' , 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12:'Dec'}
    if gamedate[4] == '-':
        dateSplit = gamedate.split('-')
        date = f'{months[int(dateSplit[1])]} {dateSplit[2]}'
    else:
        dateSplit = gamedate.split('/')
        date = f'{months[int(dateSplit[0])]} {dateSplit[1]}'

    pitchTypeDic = {'Fastball' : ('FB', 'Red', '00FF0000') , 'Four-Seam': ('FB', 'Red', '00FF0000'), 'ChangeUp' : ('CH', 'Blue', '000000FF'), 'Changeup' : ('CH', 'Blue', '000000FF'),'Slider' : ('SL', 'Green', '00339966') , 'Cutter' : ('CUT', 'Turquoise', '0040E0D0'),
     'Curveball' : ('CB', 'LimeGreen', '0000FF00') , 'Splitter' : ('SP', 'LightBlue', '0000CCFF'), 'Sinker' : ('2FB', 'Orange', '00FF6600'), 'Knuckleball' : ('KN', 'DarkGreen', '00008000'), 'Undefined' : ('N/A', 'White', '00FFFFFF'), '' : ('N/A', 'White' ,'00FFFFFF'), 'Other' : ('N/A', 'White' ,'00FFFFFF')}

    s = 1.5

    fig, ax = plt.subplots(figsize = (2.5/s, 2.95/s))
    ax.set_xlim([-30,30])
    ax.set_ylim([-30,30])
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(True)
    ax.spines['left'].set_visible(True)
    ax.spines['left'].set_position('zero')
    ax.spines['bottom'].set_position('zero')
    ax.axes.xaxis.set_visible(True)
    ax.axes.yaxis.set_visible(True)
    ax.spines['bottom'].set_linewidth(2)
    ax.spines['left'].set_linewidth(2)
    ax.spines['left'].set_color('#98012E')
    ax.spines['bottom'].set_color('#98012E')
    ax.tick_params(axis = 'both', which= 'both', bottom = False, left = False, labelleft = False, labelbottom = False)

    cur.execute(f'SELECT inducedBreak, horzBreak, {pitchType(autoTagPitchType)} FROM {tableName} WHERE pitcher = ? AND date = ?' , (pitcher, gamedate))
    allPitches = cur.fetchall()

    x = []
    y = []
    n = []
    c = []

    trackerDic = {'vertBreak' : 0, 'horzBreak' : 0, 'count' : 0}
    finalDic = {}

    sortedPitches = sorted(allPitches, key= lambda tup : tup[2])

    t = sortedPitches[0][2]
    for pitch in sortedPitches:
        try:
            if pitch[2] == t:
                trackerDic['vertBreak'] += pitch[0] 
                trackerDic['horzBreak'] += pitch[1]
                trackerDic['count'] += 1
            else:
                finalDic[t] = trackerDic
                t = pitch[2]
                trackerDic = {'vertBreak' : 0, 'horzBreak' : 0, 'count' : 0}
                trackerDic['vertBreak'] += pitch[0] 
                trackerDic['horzBreak'] += pitch[1]
                trackerDic['count'] += 1
        except:
            pass
    finalDic[t] = trackerDic

    try:
        for pitch in finalDic:        
            x.append(finalDic[pitch]['horzBreak']/finalDic[pitch]['count'])
            y.append(finalDic[pitch]['vertBreak']/finalDic[pitch]['count'])
            c.append(pitch)
    except:
        print('No pitch chart')
    
    df = pd.DataFrame({'x': x,
                   'y': y,
                   'c': c})

    groups = df.groupby('c')
    
    for name, group in groups:
        plt.scatter(group.x, group.y, s = 100/s, zorder = 3, color = pitchTypeDic[name][1], edgecolors= 'Black', linewidths = .45)
    
    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Pitch Break Graphics')
    except:
        pass
    newPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Pitch Break Graphics\{pitcher}.png'
    
    plt.savefig(newPath, bbox_inches='tight', transparent = True)
    plt.close()

    return Image(newPath)

def drawStrikezoneFigure(dbName, player, gamedate, tableName, autoTagPitchType, inning, AB):
    cur, con = openDatabase(dbName)

    months = {1 : 'Jan', 2: 'Feb' , 3: 'Mar' , 4: 'Apr' , 5 : 'May' , 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12:'Dec'}
    if gamedate[4] == '-':
        dateSplit = gamedate.split('-')
        date = f'{months[int(dateSplit[1])]} {dateSplit[2]}'
    else:
        dateSplit = gamedate.split('/')
        date = f'{months[int(dateSplit[0])]} {dateSplit[1]}'

    pitchTypeDic = {'Fastball' : ('FB', 'Red', '00FF0000') , 'Four-Seam': ('FB', 'Red', '00FF0000'), 'ChangeUp' : ('CH', 'Blue', '000000FF'), 'Changeup' : ('CH', 'Blue', '000000FF'),'Slider' : ('SL', 'Green', '00339966') , 'Cutter' : ('CUT', 'Turquoise', '0040E0D0'),
     'Curveball' : ('CB', 'LimeGreen', '0000FF00') , 'Splitter' : ('SP', 'LightBlue', '0000CCFF'), 'Sinker' : ('2FB', 'Orange', '00FF6600'), 'Knuckleball' : ('KN', 'DarkGreen', '00008000'), 'Undefined' : ('N/A', 'White', '00FFFFFF'), '' : ('N/A', 'White' ,'00FFFFFF'), 'Other' : ('N/A', 'White' ,'00FFFFFF')}
    

    s = 1.5

    fig, ax = plt.subplots(figsize = (3/s, 6/s))
    ax.set_xlim([(-1.5/s),(1.5/s)])
    ax.set_ylim([0/s,6/s])
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.axes.xaxis.set_visible(False)
    ax.axes.yaxis.set_visible(False)

    cur.execute(f'SELECT locationX, locationY, {pitchType(autoTagPitchType)} FROM {tableName} WHERE hitter = ? AND date = ? AND inning = ? AND ab = ?' , (player, gamedate, inning, AB))
    allPitches = cur.fetchall()

    x = []
    y = []
    n = []
    c = []

    num = 1
    for pitch in allPitches:
        if not pitch[0]:
            continue
        if ((pitch[0]*-1)/s) >= 1.25/s:
            x.append(1.25/s)
        elif ((pitch[0]*-1)/s) <= -1.25/s:
            x.append(-1.25/s)
        else:
            x.append((pitch[0]*-1)/s)
        if (pitch[1]/s) >= 3.7/s:
            y.append(3.7/s)
        elif (pitch[1]/s) <= 1.45/s:
            y.append(1.45/s)
        else:
            y.append(pitch[1]/s)
        n.append(num)
        c.append(pitch[2])
        num += 1

    ax.add_patch(Rectangle((-0.708333333333/s, 1.55/s), 1.416666666667/s, 1.67/s,
             edgecolor = 'black', facecolor= '#04244B',
             fill=True,
             lw = 1))

    df = pd.DataFrame({'x': x,
                   'y': y,
                   'c': c})

    groups = df.groupby('c')
    for name, group in groups:
        plt.scatter(group.x, group.y, s = 125/s, zorder = 3, color = pitchTypeDic[name][1], linewidths = .35, edgecolors = 'Black')

    for i, label in enumerate(n):
        plt.annotate(label, (x[i], y[i]), xytext = (x[i]-.045, y[i]-.056), fontsize = 8)
    
    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Strikezones')
    except:
        pass
    newPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Strikezones\{player.split(",")[0].strip()}-{date} Inning-AB: {str(inning)}-{str(AB)}.png'
    
    plt.savefig(newPath, transparent = True)
    plt.close()

    return Image(newPath)

def drawStrikezoneFigurePitcher(dbName, pitcher, gamedate, inning, tableName, autoTagPitchType):

    cur, con = openDatabase(dbName)

    months = {1 : 'Jan', 2: 'Feb' , 3: 'Mar' , 4: 'Apr' , 5 : 'May' , 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12:'Dec'}
    if gamedate[4] == '-':
        dateSplit = gamedate.split('-')
        date = f'{months[int(dateSplit[1])]} {dateSplit[2]}'
    else:
        dateSplit = gamedate.split('/')
        date = f'{months[int(dateSplit[0])]} {dateSplit[1]}'

    pitchTypeDic = {'Fastball' : ('FB', 'Red', '00FF0000') , 'Four-Seam': ('FB', 'Red', '00FF0000'), 'ChangeUp' : ('CH', 'Blue', '000000FF'), 'Changeup' : ('CH', 'Blue', '000000FF'),'Slider' : ('SL', 'Green', '00339966') , 'Cutter' : ('CUT', 'Turquoise', '0040E0D0'),
     'Curveball' : ('CB', 'LimeGreen', '0000FF00') , 'Splitter' : ('SP', 'LightBlue', '0000CCFF'), 'Sinker' : ('2FB', 'Orange', '00FF6600'), 'Knuckleball' : ('KN', 'DarkGreen', '00008000'), 'Undefined' : ('N/A', 'White', '00FFFFFF'), '' : ('N/A', 'White' ,'00FFFFFF'), 'Other' : ('N/A', 'White' ,'00FFFFFF')}

    s = 1.5

    fig, ax = plt.subplots(figsize = (3/s, 6/s))
    ax.set_xlim([(-1.5/s),(1.5/s)])
    ax.set_ylim([0/s,6/s])
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.axes.xaxis.set_visible(False)
    ax.axes.yaxis.set_visible(False)

    cur.execute(f'SELECT locationX, locationY, {pitchType(autoTagPitchType)} FROM {tableName} WHERE pitcher = ? AND date = ? AND inning = ?' , (pitcher, gamedate, inning))
    allPitches = cur.fetchall()

    x = []
    y = []
    n = []
    c = []

    num = 1
    for pitch in allPitches:
        if not pitch[0]:
            continue
        if ((pitch[0])/s) >= 1.25/s:
            x.append(1.25/s)
        elif ((pitch[0])/s) <= -1.25/s:
            x.append(-1.25/s)
        else:
            x.append((pitch[0])/s)
        if (pitch[1]/s) >= 3.9/s:
            y.append(3.9/s)
        elif (pitch[1]/s) <= 1.25/s:
            y.append(1.25/s)
        else:
            y.append(pitch[1]/s)
        n.append(num)
        c.append(pitch[2])
        num += 1

    ax.add_patch(Rectangle((-0.708333333333/s, 1.55/s), 1.416666666667/s, 1.67/s,
             edgecolor = 'black', facecolor= '#04244B',
             fill=True,
             lw = 1)),

    df = pd.DataFrame({'x': x,
                   'y': y,
                   'c': c})

    groups = df.groupby('c')
    for name, group in groups:
        plt.scatter(group.x, group.y, s = 100/s, zorder = 3, color = pitchTypeDic[name][1], edgecolors= 'Black', linewidths = .45)
    
    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Strikezone Graphics')
    except:
        pass
    newPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Strikezone Graphics\{pitcher}-{inning}.png'
    
    plt.savefig(newPath, transparent = True)
    plt.close()

    return Image(newPath)

def getPitcherPitches(dbName, pitcher, date, tableName, autoTagPitchType, inning = ''):
    cur, con = openDatabase(dbName)
    vars = f'pitchVelo,spinRate,{pitchType(autoTagPitchType)},pitchResult,exitVelo,locationX,locationY,inning,outcome,inducedBreak,vertBreak,horzBreak'

    if inning:
        cur.execute(f'SELECT {vars} FROM {tableName} WHERE pitcher = ? AND date = ? AND inning = ?' , (pitcher, date, inning))
        allPitches = cur.fetchall()
    else:
        cur.execute(f'SELECT {vars} FROM {tableName} WHERE pitcher = ? AND date = ?' , (pitcher, date))
        allPitches = cur.fetchall()

    return allPitches

def getInningsPitched(dbName, pitcher, date, tableName, autoTagPitchType):
    allPitches = getPitcherPitches(dbName, pitcher, date, tableName, autoTagPitchType, inning = '')
    sortedAllPitches = sorted(allPitches, key= lambda tup : tup[7])

    try:
        i = sortedAllPitches[0][7]
    except IndexError:
        print(f'Exception pulling pitches from database: {pitcher}')
        i = 1
    innings = [i]
    for pitch in sortedAllPitches:
        if pitch[7] != i:
            i = pitch[7]
            innings.append(pitch[7])
    return innings

def getPitchTypeInfo(dbName, pitcher, date, tableName, autoTagPitchType, inning = ''):

    pitchTypeInfo = {}
    allPitches = getPitcherPitches(dbName, pitcher, date, tableName, autoTagPitchType, inning)
    sortedAllPitches = sorted(allPitches, key= lambda tup : tup[7])
    pitchTypeInfo['all'] = {'maxVelo' : 0, 'velo' : 0, 'spin' : 0, 'veloSpinCount' : 0, 'exitVeloCount': 0, 'usageCount' : 0, 'strikeCount' : 0, 'hardHitCount' : 0, 'chaseCount' : 0, 'whiffCount' : 0, 'inducedBreak' : 0, 'vertBreak' : 0, 'horzBreak' : 0}
    
    for pitch in sortedAllPitches:
        ball = False
        pitchTypeInfo['all']['usageCount'] += 1
        if pitch[2] not in pitchTypeInfo:
            pitchTypeInfo[pitch[2]] = {'maxVelo' : 0,'velo' : 0, 'spin' : 0, 'veloSpinCount' : 0, 'exitVeloCount': 0, 'usageCount' : 1, 'strikeCount' : 0, 'hardHitCount' : 0, 'chaseCount' : 0, 'whiffCount' : 0, 'inducedBreak' : 0, 'vertBreak' : 0, 'horzBreak' : 0}
            if pitch[0] and pitch[1]:
                if pitch[0] > pitchTypeInfo[pitch[2]]['maxVelo']:
                    pitchTypeInfo[pitch[2]]['maxVelo'] = pitch[0]
                pitchTypeInfo[pitch[2]]['velo'] += pitch[0]
                pitchTypeInfo[pitch[2]]['spin'] += pitch[1]
                pitchTypeInfo[pitch[2]]['veloSpinCount'] += 1
                pitchTypeInfo['all']['veloSpinCount'] += 1

                try:
                    pitchTypeInfo[pitch[2]]['inducedBreak'] += pitch[9]
                    pitchTypeInfo[pitch[2]]['vertBreak'] += pitch[10]
                    pitchTypeInfo[pitch[2]]['horzBreak'] += pitch[11]
                except TypeError:
                    print(f'Error with reading break: {pitcher}')

            if pitch[3][0] == 'B':
                pass
            else:
                pitchTypeInfo[pitch[2]]['strikeCount'] += 1
                pitchTypeInfo['all']['strikeCount'] += 1

            if pitch[4]:
                pitchTypeInfo[pitch[2]]['exitVeloCount'] += 1
                pitchTypeInfo['all']['exitVeloCount'] += 1
                if pitch[4] >= 90:
                    pitchTypeInfo[pitch[2]]['hardHitCount'] += 1
                    pitchTypeInfo['all']['hardHitCount'] += 1

            if pitch[5]:
                if pitch[5] < -0.708333333333333 or pitch[5] > 0.7083333333333333:
                    ball = True
                if pitch[6] < 1.55 or pitch[6] > (1.55+1.67):
                    ball = True
                if ball and pitch[3] == ('StrikeSwinging' or 'InPlay' or 'FoulBall'):
                    pitchTypeInfo[pitch[2]]['chaseCount'] += 1
                    pitchTypeInfo['all']['chaseCount'] += 1

            if pitch[3] == 'StrikeSwinging':
                pitchTypeInfo[pitch[2]]['whiffCount'] += 1
                pitchTypeInfo['all']['whiffCount'] += 1
        else:
            if pitch[0] and pitch[1]:
                if pitch[0] > pitchTypeInfo[pitch[2]]['maxVelo']:
                    pitchTypeInfo[pitch[2]]['maxVelo'] = pitch[0]
                pitchTypeInfo[pitch[2]]['velo'] += pitch[0]
                pitchTypeInfo[pitch[2]]['spin'] += pitch[1]
                pitchTypeInfo[pitch[2]]['veloSpinCount'] += 1
                pitchTypeInfo['all']['veloSpinCount'] += 1

                try:
                    pitchTypeInfo[pitch[2]]['inducedBreak'] += pitch[9]
                    pitchTypeInfo[pitch[2]]['vertBreak'] += pitch[10]
                    pitchTypeInfo[pitch[2]]['horzBreak'] += pitch[11]
                except TypeError:
                    print(f'Error with reading break: {pitcher}')

            if pitch[3][0] == 'B':
                pass
            else:
                pitchTypeInfo[pitch[2]]['strikeCount'] += 1
                pitchTypeInfo['all']['strikeCount'] += 1

            pitchTypeInfo[pitch[2]]['usageCount'] += 1

            if pitch[4]:
                pitchTypeInfo[pitch[2]]['exitVeloCount'] += 1
                pitchTypeInfo['all']['exitVeloCount'] += 1
                if pitch[4] >= 90:
                    pitchTypeInfo[pitch[2]]['hardHitCount'] += 1
                    pitchTypeInfo['all']['hardHitCount'] += 1
                    

            if pitch[5]:
                if pitch[5] < -0.708333333333333 or pitch[5] > 0.7083333333333333:
                    ball = True
                if pitch[6] < 1.55 or pitch[6] > (1.55+1.67):
                    ball = True
                if ball and pitch[3] == ('StrikeSwinging' or 'InPlay' or 'FoulBall'):
                    pitchTypeInfo[pitch[2]]['chaseCount'] += 1
                    pitchTypeInfo['all']['chaseCount'] += 1

            if pitch[3] == 'StrikeSwinging':
                pitchTypeInfo[pitch[2]]['whiffCount'] += 1
                pitchTypeInfo['all']['whiffCount'] += 1

    return pitchTypeInfo

def sortPitchTypeInfo(pitchTypeInfo):
    totalUsage = 0
    pitches = []
    for pitch, dic in pitchTypeInfo.items():
        if pitch != 'all':
            totalUsage += dic['usageCount']
        pitches.append((pitch, dic['usageCount']))

    return sorted(pitches, key= lambda tup: tup[1], reverse = True), totalUsage

def addCellColor(sheet, count):
    colorDic = {'red' : '00FF0000', 'green' : '00339966', 'yellow' : '00FFFF00'}
    
    color = 'yellow'
    if sheet[f'L{count}'].internal_value > .65:
        color = 'green'
    elif sheet[f'L{count}'].internal_value < .60:
        color = 'red'
    sheet[f'L{count}'].fill = PatternFill('solid', fgColor= colorDic[color])
    color = 'yellow'
    if sheet[f'M{count}'].internal_value < .30:
        color = 'green'
    elif sheet[f'M{count}'].internal_value > .4:
        color = 'red'
    sheet[f'M{count}'].fill = PatternFill('solid', fgColor= colorDic[color])
    color = 'yellow'
    if sheet[f'N{count}'].internal_value > .25:
        color = 'green'
    elif sheet[f'N{count}'].internal_value < .15:
        color = 'red'
    sheet[f'N{count}'].fill = PatternFill('solid', fgColor= colorDic[color])
    color = 'yellow'
    if sheet[f'O{count}'].internal_value > .25:
        color = 'green'
    elif sheet[f'O{count}'].internal_value < .15:
        color = 'red'
    sheet[f'O{count}'].fill = PatternFill('solid', fgColor= colorDic[color])
    
def addStrikeZone(imgCount, dbName, player, gamedate, allPitches, sheet, i, tableName, autoTagPitchType):
    img = drawStrikezoneFigure(dbName, player, gamedate, tableName, autoTagPitchType, inning=allPitches[i][7], AB=allPitches[i][8])
    sheet.add_image(img, f'B{imgCount}')
    if imgCount == 1:
        imgCount += 9
    else:
        imgCount += 8
    return imgCount

def getVelo(cur, tableName, pitcher):

    cur.execute(f'SELECT pitchVelo FROM {tableName} WHERE pitcher = ? AND taggedPitch = ?' , (pitcher, 'Fastball'))
    tups = cur.fetchall()

    if len(tups) < 3:
        cur.execute(f'SELECT pitchVelo FROM {tableName} WHERE pitcher = ? AND autoPitch = ?' , (pitcher, 'Four-Seam'))
        tups = cur.fetchall()

    slow = 200
    fast = 0
    try:
        for tup in tups:
            if float(tup[0]) < slow:
                slow = float(tup[0])
            if float(tup[0]) > fast:
                fast = float(tup[0])
        velo = f'{int(slow)}-{int(fast)}'
    except:
        velo = 'Error'

    try:
        for tup in tups:
            if float(tup[0]) < slow:
                slow = float(tup[0])
            if float(tup[0]) > fast:
                fast = float(tup[0])
        velo = f'{int(slow)}-{int(fast)}'
    except:
        velo = 'Error'

    if slow == 200 or fast == 0:
        velo = 'Error'

    return velo

def getMix(cur, tableName, pitcher):

    cur.execute(f'SELECT {pitchType(autoTagPitchType)} FROM {tableName} WHERE pitcher = ?' , (pitcher,))
    tup = cur.fetchall()

    pitchDic = {'Fastball' : ('FB', 'Red', '00FF0000') , 'Four-Seam': ('FB', 'Red', '00FF0000'), 'ChangeUp' : ('CH', 'Blue', '000000FF'), 'Changeup' : ('CH', 'Blue', '000000FF'),'Slider' : ('SL', 'Green', '00339966') , 'Cutter' : ('CUT', 'Turquoise', '0040E0D0'),
     'Curveball' : ('CB', 'LimeGreen', '0000FF00') , 'Splitter' : ('SP', 'LightBlue', '0000CCFF'), 'Sinker' : ('2FB', 'Orange', '00FF6600'), 'Knuckleball' : ('KN', 'DarkGreen', '00008000'), 'Undefined' : ('??', 'White', '00FFFFFF'), '' : ('??', 'White' ,'00FFFFFF'), 'Other' : ('??', 'White' ,'00FFFFFF')}

    pitchList = []
    pitchString = ''
    for pitch in tup:
        if pitch[0] not in pitchList:
            pitchList.append(pitch[0])
            pitchString += (pitchDic[pitch[0]][0] + '-')

    return pitchString.rstrip('-')

def getSynergy(pitcher, hitter, tb, inning, outs):
    
    syn = f'Access Synergy Video: {pitcher.split()[1]}, {pitcher.split()[0]} VS {hitter.split()[0]} {hitter.split()[1][0]}. {tb[0]}{inning}, {outs} Out'

    return syn

def addIColumn(allPitches, i, sheet, count, abCount, cur, tableName):
    inning = allPitches[i][7]
    outs = allPitches[i][2]

    pitcher = f'{allPitches[i][1].split()[1][0]}. {allPitches[i][1].split()[0].strip(",")}'
    throws = allPitches[i][9]
    velo = getVelo(cur, tableName, allPitches[i][1])
    mix = getMix(cur, tableName, allPitches[i][1])

    hitter = allPitches[i][11]
    tb = allPitches[i][10]

    syn = getSynergy(pitcher, hitter, tb, inning, outs)

    count =  8 * abCount + 2
    if count >= 50:
        count += 7

    sheet[f'I{count}'] = inning
    count += 1
    sheet[f'I{count}'] = outs
    count += 2

    sheet[f'I{count}'] = pitcher
    count += 1
    sheet[f'I{count}'] = throws
    count += 1
    sheet[f'I{count}'] = f'{velo} MPH'
    count += 1
    sheet[f'I{count}'] = mix
    count += 1
    sheet[f'H{count}'] = syn
    
def addQAB(sheet, abCount, count):
    count -= 6
    qab = 'No'
    #3 pitches after 2 ks
    strikeCount = 0
    post2kCount = 0
    hitList = ['Single', 'Double', 'Triple', 'Home run', 'Home Run', 'Walk', 'Hit by pitch', 'HBP', 'SAC', 'Sac fly', 'Sac', 'Sac Fly', 'Sacrifice']
    strikeList = ['Strike', 'Swing', 'Foul']
    endList = ['Strike', 'Swing', 'In Play']
    extra = 0

    if abCount >= 6:
        extra = 7
    for num in range(0, 8):
        if strikeCount == 2:
            post2kCount += 1
        elif strikeCount == 3:
            if sheet[f'G{10+num+((abCount-1)*8)+extra}'].internal_value in endList:
                break
        if sheet[f'G{10+num+((abCount-1)*8)+extra}'].internal_value in strikeList:
            strikeCount += 1
        elif sheet[f'G{10+num+((abCount-1)*8)+extra}'].internal_value in endList:
            break

    if post2kCount >= 3:
        qab = 'Yes'
    #7+ pitches
    if sheet[f'G{8+(abCount*8)+extra}'].internal_value:
        qab = 'Yes'
    #base hit
    if sheet[f'L{7+(abCount*8)+extra}'].internal_value in hitList:
        qab = 'Yes'
    #Barrelled
    try:
        if float(sheet[f'L{2+(abCount*8)+extra}'].internal_value[:-4]) > 100:
            qab = 'Yes'
        elif float(sheet[f'L{2+(abCount*8)+extra}'].internal_value[:-4]) > 90:
            if float(sheet[f'L{4+(abCount*8)+extra}'].internal_value[:-2]) >= 8 and float(sheet[f'L{4+(abCount*8)+extra}'].internal_value[:-2]) <= 32.5:
                qab = 'Yes'
    except:
        pass
    count += 6
    sheet[f'L{count}'] = qab

def addLColumn(abCount, allPitches, sheet, i, count):
    hitTypeDic = {'LineDrive' : 'Line Drive', 'GroundBall' : 'Ground Ball', 'FlyBall' : 'Fly Ball', 'Popup' : 'Popup'}
    resultDic = {'Single' : 'Single', 'Double' : 'Double', 'Triple' : 'Triple', 'HomeRun' : 'Home Run', 'FieldersChoice' : 'Fielders Choice', 'Out' : 'Out', 'Sacrifice' : 'Sacrifice', 'Error' : 'Error'}

    count =  8 * abCount + 2
    if count >= 50:
        count += 7

    if allPitches[i][3] == 'NA':
        sheet[f'L{count}'] = "--"
    elif not allPitches[i][3]:
        sheet[f'L{count}'] = "--"
    else:
        sheet[f'L{count}'] = f'{round(allPitches[i][3], 2)} MPH'
    count += 2
    if allPitches[i][4] == 'NA':
        sheet[f'L{count}'] = "--"
    elif not allPitches[i][4]:
        sheet[f'L{count}'] = "--"
    else:
        sheet[f'L{count}'] = f'{round(allPitches[i][4], 2)}°'

    count += 2
    try:
        sheet[f'L{count}'] = hitTypeDic[allPitches[i][5]]
    except:
        sheet[f'L{count}'] = "--"
    
    count += 1
    
    try:
        sheet[f'L{count}'] = resultDic[allPitches[i][6]]
    except:
        if allPitches[i][13] == 'Strikeout':
            sheet[f'L{count}'] = 'Strikeout'
        elif allPitches[i][13] == 'Walk':
            sheet[f'L{count}'] = 'Walk'
        elif allPitches[i][12] == 'HitByPitch':
            sheet[f'L{count}'] = 'HBP'
    
    count += 1
    addQAB(sheet, abCount, count)

def getBestSequence(dbName, pitcher, date, tableName, autoTagPitchType, inning = ''):
    pitches = getPitcherPitches(dbName, pitcher, date, tableName, autoTagPitchType, inning)
    pitchTypeDic = {'Fastball' : ('FB', 'Red', '00FF0000') , 'Four-Seam': ('FB', 'Red', '00FF0000'), 'ChangeUp' : ('CH', 'Blue', '000000FF'), 'Changeup' : ('CH', 'Blue', '000000FF'),'Slider' : ('SL', 'Green', '00339966') , 'Cutter' : ('FC', 'Turquoise', '0040E0D0'),
     'Curveball' : ('CB', 'LimeGreen', '0000FF00') , 'Splitter' : ('SP', 'LightBlue', '0000CCFF'), 'Sinker' : ('SK', 'Orange', '00FF6600'), 'Knuckleball' : ('KN', 'DarkGreen', '00008000'), 'Undefined' : ('NA', 'White', '00FFFFFF'), '' : ('NA', 'White' ,'00FFFFFF'), 'Other' : ('NA', 'White' ,'00FFFFFF')}

    last3 = ''
    hitList = ['Single', 'Double', 'Triple', 'HomeRun']
    outList = ['Out', 'FieldersChoice', 'Error']
    rankings = {}
    for pitch in pitches:
        outcome = pitch[8]
        pitchType = pitch[2]
        
        if outcome == 'Undefined':
            if len(last3) < 9:
                last3 += pitchTypeDic[pitchType][0] + '-'
            elif len(last3) == 9:
                last3 = last3[3:9]
                last3 += pitchTypeDic[pitchType][0] + '-'
        elif outcome in outList:
            if len(last3) < 9:
                last3 += pitchTypeDic[pitchType][0] + '-'
            elif len(last3) == 9:
                last3 = last3[3:9]
                last3 += pitchTypeDic[pitchType][0] + '-'
                if last3 in rankings:
                    rankings[last3] += 1
                else:
                    rankings[last3] = 1
            last3 = ''
        elif outcome in hitList:
            if len(last3) < 9:
                last3 += pitchTypeDic[pitchType][0] + '-'
            elif len(last3) == 9:
                last3 = last3[3:9]
                last3 += pitchTypeDic[pitchType][0] + '-'
                if last3 in rankings:
                    rankings[last3] -= 1
                else:
                    rankings[last3] = -1
            last3 = ''
        elif outcome == 'Sacrifice':
            last3 = ''
        else:
            print(f'New outcome: {outcome}')
            last3 = ''

    return sorted(rankings.keys(), key= lambda last3 : rankings[last3], reverse = True)

def writeBestSequences(sheet, count, dbName, pitcher, date, tableName, autoTagPitchType, inning = ''):
    list = getBestSequence(dbName, pitcher, date, tableName, autoTagPitchType, inning)

    try:
        sheet[f'M{count}'] = f'{list[0].rstrip("-")} & {list[1].rstrip("-")}'
    except:
        try:
            sheet[f'M{count}'] = f'{list[0].rstrip("-")}'
        except:
            sheet[f'M{count}'] = 'No Patterns Found'
            print(f'No good sequences found: {pitcher}')

def getOverallStats(dbName, tableName, pitcher):
    
    cur, con = openDatabase(dbName)
    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ?' , ('Out', pitcher))
    outs = len(cur.fetchall())
    cur.execute(f'SELECT hitter FROM {tableName} WHERE KorBB = ? AND pitcher = ?' , ('Strikeout', pitcher))
    strikeouts = len(cur.fetchall())

    innings = (outs + strikeouts) / 3

    if str(round(innings, 1))[-1] == '3':
        innings = int(str(innings)[:1]) + .1
    elif str(round(innings, 1))[-1] == '6':
        innings = int(str(innings)[:1]) + .2

    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ?' , ('Single', pitcher))
    singles = len(cur.fetchall())
    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ?' , ('Double', pitcher))
    doubles = len(cur.fetchall())
    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ?' , ('Triple', pitcher))
    triples = len(cur.fetchall())
    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ?' , ('HomeRun', pitcher))
    HRs = len(cur.fetchall())

    hits = singles + doubles + triples + HRs

    cur.execute(f'SELECT hitter FROM {tableName} WHERE pitcher = ? AND runsScored >= ?', (pitcher, 1))
    runs = len(cur.fetchall())

    cur.execute(f'SELECT hitter FROM {tableName} WHERE KorBB = ? AND pitcher = ?' ,('Walk', pitcher))
    walks = len(cur.fetchall())

    return f'{strikeouts} K, {runs} R, {hits} H, {walks} BB'

def getInningStats(dbName, tableName, pitcher, inning):

    cur, con = openDatabase(dbName)
    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ? AND inning = ?' , ('Out', pitcher, inning))
    outs = len(cur.fetchall())
    cur.execute(f'SELECT hitter FROM {tableName} WHERE KorBB = ? AND pitcher = ? AND inning = ?' , ('Strikeout', pitcher, inning))
    strikeouts = len(cur.fetchall())

    innings = (outs + strikeouts) / 3

    if str(round(innings, 1))[-1] == '3':
        innings = int(str(innings)[:1]) + .1
    elif str(round(innings, 1))[-1] == '6':
        innings = int(str(innings)[:1]) + .2

    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ? AND inning = ?' , ('Single', pitcher, inning))
    singles = len(cur.fetchall())
    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ? AND inning = ?' , ('Double', pitcher, inning))
    doubles = len(cur.fetchall())
    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ? AND inning = ?' , ('Triple', pitcher, inning))
    triples = len(cur.fetchall())
    cur.execute(f'SELECT hitter FROM {tableName} WHERE outcome = ? AND pitcher = ? AND inning = ?' , ('HomeRun', pitcher, inning))
    HRs = len(cur.fetchall())

    hits = singles + doubles + triples + HRs

    cur.execute(f'SELECT hitter FROM {tableName} WHERE inning = ? AND pitcher = ? AND runsScored >= ?', (inning, pitcher, 1))
    runs = len(cur.fetchall())

    cur.execute(f'SELECT hitter FROM {tableName} WHERE KorBB = ? AND pitcher = ? AND inning = ?', ('Walk', pitcher, inning))
    walks = len(cur.fetchall())

    return f'{strikeouts} K, {runs} R, {hits} H, {walks} BB'

def writeOverall(sheet, dbName, pitcher, tableName, date, autoTagPitchType):
    pitchTypeInfo = getPitchTypeInfo(dbName, pitcher, date, tableName, autoTagPitchType)
    sortPitches, totalUsage = sortPitchTypeInfo(pitchTypeInfo)
    pitchTypeDic = {'Fastball' : ('FB', 'Red', '00FF0000') , 'Four-Seam': ('FB', 'Red', '00FF0000'), 'ChangeUp' : ('CH', 'Blue', '000000FF'), 'Changeup' : ('CH', 'Blue', '000000FF'),'Slider' : ('SL', 'Green', '00339966') , 'Cutter' : ('CUT', 'Turquoise', '0040E0D0'),
     'Curveball' : ('CB', 'LimeGreen', '0000FF00') , 'Splitter' : ('SP', 'LightBlue', '0000CCFF'), 'Sinker' : ('2FB', 'Orange', '00FF6600'), 'Knuckleball' : ('KN', 'DarkGreen', '00008000'), 'Undefined' : ('OTH', 'White', '00FFFFFF'), '' : ('OTH', 'White' ,'00FFFFFF'), 'Other' : ('OTH', 'White' ,'00FFFFFF')}

    count = 10
    for pitch, usage in sortPitches[1:]:
        sheet[f'F{count}'] = pitchTypeDic[pitch][0]
        sheet[f'F{count}'].font = Font(color = pitchTypeDic[pitch][2], bold= True)
        sheet[f'G{count}'] = pitchTypeInfo[pitch]['maxVelo']
        try:
            sheet[f'H{count}'] = pitchTypeInfo[pitch]['velo']/pitchTypeInfo[pitch]['veloSpinCount']
            sheet[f'I{count}'] = pitchTypeInfo[pitch]['inducedBreak']/pitchTypeInfo[pitch]['veloSpinCount']
            sheet[f'J{count}'] = pitchTypeInfo[pitch]['horzBreak']/pitchTypeInfo[pitch]['veloSpinCount']
        except ZeroDivisionError:
            print(f'Exception occured. No velo recorded for pitch: {pitch}, {pitcher}')
        sheet[f'K{count}'] = pitchTypeInfo[pitch]['usageCount']/totalUsage
        sheet[f'L{count}'] = pitchTypeInfo[pitch]['strikeCount']/pitchTypeInfo[pitch]['usageCount']
        try:
            sheet[f'M{count}'] = pitchTypeInfo[pitch]['hardHitCount']/pitchTypeInfo[pitch]['exitVeloCount']
        except ZeroDivisionError:
            sheet[f'M{count}'] = 0
        sheet[f'N{count}'] = pitchTypeInfo[pitch]['chaseCount']/pitchTypeInfo[pitch]['usageCount']
        sheet[f'O{count}'] = pitchTypeInfo[pitch]['whiffCount']/pitchTypeInfo[pitch]['usageCount']
        addCellColor(sheet, count)
        count += 2
        if count == 18:
            break
    count = 18
    sheet[f'L{count}'] = pitchTypeInfo['all']['strikeCount']/totalUsage
    try:
        sheet[f'M{count}'] = pitchTypeInfo['all']['hardHitCount']/pitchTypeInfo['all']['exitVeloCount']
    except ZeroDivisionError:
        sheet[f'M{count}'] = 0
    sheet[f'N{count}'] = pitchTypeInfo['all']['chaseCount']/totalUsage
    sheet[f'O{count}'] = pitchTypeInfo['all']['whiffCount']/totalUsage
    addCellColor(sheet, count)
    count += 2
    writeBestSequences(sheet, count, dbName, pitcher, date, tableName, autoTagPitchType)
    stats = getOverallStats(dbName, tableName, pitcher) 
    sheet[f'F{count}'] = stats

    try:
        img = drawPitchBreakChart(date, pitcher, autoTagPitchType, dbName, tableName)
    except:
        print(f'No Image Drawn: {pitcher}')
    try:
        sheet.add_image(img,'C10')
    except:
        pass

def writeInnings(sheet, dbName, pitcher, inning, inningCount, tableName, date, autoTagPitchType):
    pitchTypeInfo = getPitchTypeInfo(dbName, pitcher, date, tableName, autoTagPitchType, inning)
    sortPitches, totalUsage = sortPitchTypeInfo(pitchTypeInfo)
    pitchTypeDic = {'Fastball' : ('FB', 'Red', '00FF0000') , 'Four-Seam': ('FB', 'Red', '00FF0000'), 'ChangeUp' : ('CH', 'Blue', '000000FF'), 'Changeup' : ('CH', 'Blue', '000000FF'),'Slider' : ('SL', 'Green', '00339966') , 'Cutter' : ('CUT', 'Turquoise', '0040E0D0'),
     'Curveball' : ('CB', 'LimeGreen', '0000FF00') , 'Splitter' : ('SP', 'LightBlue', '0000CCFF'), 'Sinker' : ('2FB', 'Orange', '00FF6600'), 'Knuckleball' : ('KN', 'DarkGreen', '00008000'), 'Undefined' : ('N/A', 'White', '00FFFFFF'), '' : ('N/A', 'White' ,'00FFFFFF'), 'Other' : ('N/A', 'White' ,'00FFFFFF')}

    count = (inningCount+1)*11
    if count >= 77:
        count += 1

    for pitch, usage in sortPitches[1:]:
        sheet[f'F{count}'] = pitchTypeDic[pitch][0]
        sheet[f'F{count}'].font = Font(color = pitchTypeDic[pitch][2], bold= True)
        sheet[f'G{count}'] = pitchTypeInfo[pitch]['maxVelo']
        try:
            sheet[f'H{count}'] = pitchTypeInfo[pitch]['velo']/pitchTypeInfo[pitch]['veloSpinCount']
            sheet[f'I{count}'] = pitchTypeInfo[pitch]['inducedBreak']/pitchTypeInfo[pitch]['veloSpinCount']
            sheet[f'J{count}'] = pitchTypeInfo[pitch]['horzBreak']/pitchTypeInfo[pitch]['veloSpinCount']
        except ZeroDivisionError:
            print(f'Exception occured. No velo or spin recorded for pitch: {pitch} {pitcher} in {inning} inning')
        sheet[f'K{count}'] = pitchTypeInfo[pitch]['usageCount']/totalUsage
        sheet[f'L{count}'] = pitchTypeInfo[pitch]['strikeCount']/pitchTypeInfo[pitch]['usageCount']
        try:
            sheet[f'M{count}'] = pitchTypeInfo[pitch]['hardHitCount']/pitchTypeInfo[pitch]['exitVeloCount']
        except ZeroDivisionError:
            sheet[f'M{count}'] = 0
        sheet[f'N{count}'] = pitchTypeInfo[pitch]['chaseCount']/pitchTypeInfo[pitch]['usageCount']
        sheet[f'O{count}'] = pitchTypeInfo[pitch]['whiffCount']/pitchTypeInfo[pitch]['usageCount']
        addCellColor(sheet, count)
        count += 2
        if count == (inningCount+1)*11 + 8:
            break

    count = (inningCount+1)*11 + 8
    if count >= 77:
        count += 1
    
    sheet[f'L{count}'] = pitchTypeInfo['all']['strikeCount']/totalUsage
    
    try:
            sheet[f'M{count}'] = pitchTypeInfo['all']['hardHitCount']/pitchTypeInfo['all']['exitVeloCount']
    except ZeroDivisionError:
            sheet[f'M{count}'] = 0
    sheet[f'N{count}'] = pitchTypeInfo['all']['chaseCount']/totalUsage
    sheet[f'O{count}'] = pitchTypeInfo['all']['whiffCount']/totalUsage
    addCellColor(sheet, count)
    count += 2
    writeBestSequences(sheet, count, dbName, pitcher, date, tableName, autoTagPitchType, inning)
    stats = getInningStats(dbName, tableName, pitcher, inning)
    sheet[f'F{count}'] = stats

def addStrikezoneFigure(dbName, pitcher, date, inning, imgCount, sheet, tableName, autoTagPitchType):
    img = drawStrikezoneFigurePitcher(dbName, pitcher, date, inning, tableName, autoTagPitchType)
    sheet.add_image(img, f'B{imgCount}')

def writeHeaderPitcher(sheet, pitcher, date, opponent):
    sheet[f'C3'] = pitcher
    sheet[f'C5'] = date
    sheet[f'C7'] = f'v {opponent}'

def writeAB(abVars, player, gamedate, sheet, cur, dbName, tableName, autoTagPitchType):

    cur.execute(f'SELECT {abVars} FROM {tableName} WHERE hitter = ? AND date = ?' , (player, gamedate))
    allPitches = cur.fetchall()

    count = 10 
    abCount = 1 
    num = allPitches[0][0] 
    imgCount = 1
    SynCount = 17
    for i in range(len(allPitches)):
        if allPitches[i][0] == num:
            num += 1
        else:
            if count >= 50:
                count += 7
                imgCount += 7
                SynCount += 7
            i -= 1

            imgCount = addStrikeZone(imgCount, dbName, player, gamedate, allPitches, sheet, i, tableName, autoTagPitchType)
            addIColumn(allPitches, i, sheet, count, abCount, cur, tableName)
            addLColumn(abCount, allPitches, sheet, i, count)

            SynCount += 8
            abCount += 1
            count =  8 * abCount + 2
            num = allPitches[i+1][0] + 1

    if count >= 50:
        count += 7
        imgCount += 7
        SynCount += 7
    
    addStrikeZone(imgCount, dbName, player, gamedate, allPitches, sheet, -1, tableName, autoTagPitchType)
    addIColumn(allPitches, -1, sheet, count, abCount, cur, tableName)
    addLColumn(abCount, allPitches, sheet, -1, count)

def writeHeader(headerVars, player, sheet, cur, opponent, tableName):

    cur.execute(f'SELECT {headerVars} FROM {tableName} WHERE hitter = ?' , (player,))
    headers_ = cur.fetchone()
    
    try:
        sheet['C7'] = f'vs {opponent}'
        sheet['C3'] = headers_[1]
        sheet['C5'] = headers_[2]
        sheet['C54'] = f'vs {opponent}'
        sheet['C50'] = headers_[1]
        sheet['C52'] = headers_[2]
    except:
        pass

def printWorkbookHitter(path, player):

    if os.path.isfile(path):
        try:
            os.startfile(path, 'print')
            print(f'Printing report: {player.split()[1].rstrip(",")} {player.split()[0].rstrip(",")}')
        except:
            print(f'Exception occured: {player}')
    else:
        print(f'Error with file: {player}')
    
    time.sleep(3)
    keyboard.write(f"{player.split()[1].rstrip(',')} {player.split()[0].rstrip(',')} - Hitting")
    keyboard.press_and_release('enter')
    time.sleep(1)

def printWorkbookPitcher(path, player):

    if os.path.isfile(path):
        try:
            os.startfile(path, 'print')
            print(f'Printing report: {player.split()[1].rstrip(",")} {player.split()[0].rstrip(",")}')
        except:
            print(f'Exception occured: {player}')
    else:
        print(f'Error with file: {player}')
    
    time.sleep(3)
    keyboard.write(f"{player.split()[1].rstrip(',')} {player.split()[0].rstrip(',')} - Pitching")
    keyboard.press_and_release('enter')
    time.sleep(1)

def sendHitterWorkbooks(players):
    
    keyboard.press_and_release('c')

    time.sleep(2)
    keyboard.write('Brooke')
    time.sleep(1)
    keyboard.press_and_release('enter')

    time.sleep(1)

    keyboard.press_and_release('ctrl + shift + c')
    time.sleep(1)
    keyboard.write('Ed')
    time.sleep(1)

    keyboard.press_and_release('enter')
    keyboard.press_and_release('tab')

    time.sleep(1)
    keyboard.write('Postgame Hitting Reports')

    time.sleep(1)
    keyboard.press_and_release('tab')
    keyboard.write('Here is tonights reports.\n\n\nRyley')

    time.sleep(1)

    keyboard.press_and_release('tab')
    keyboard.press_and_release('tab')
    keyboard.press_and_release('tab')
    keyboard.press_and_release('enter')
    time.sleep(1)

    for player in players:
        keyboard.write(f'"{player} - Hitting"')

    time.sleep(1)
    keyboard.press_and_release('enter')
    time.sleep(10)
    keyboard.press_and_release('ctrl + enter')
    time.sleep(3)

def sendPitcherWorkbooks(pitchers):

    keyboard.press_and_release('c')

    time.sleep(2)
    keyboard.write('Brooke')
    time.sleep(1)
    keyboard.press_and_release('enter')

    time.sleep(1)

    keyboard.press_and_release('ctrl + shift + c')
    time.sleep(1)
    keyboard.write('Beau')
    time.sleep(1)

    keyboard.press_and_release('enter')

    time.sleep(1)
    keyboard.write('Yo')
    time.sleep(1)

    keyboard.press_and_release('enter')
    keyboard.press_and_release('tab')

    time.sleep(1)
    keyboard.write('Postgame Pitching Reports')

    time.sleep(1)
    keyboard.press_and_release('tab')
    keyboard.write('Here is tonights report.\n\n\nRyley')

    time.sleep(1)

    keyboard.press_and_release('tab')
    keyboard.press_and_release('tab')
    keyboard.press_and_release('tab')
    keyboard.press_and_release('enter')
    time.sleep(1)

    for pitcher in pitchers:
        keyboard.write(f'"{pitcher} - Pitching"')
        time.sleep(.25)

    time.sleep(1)
    keyboard.press_and_release('enter')
    time.sleep(10)
    keyboard.press_and_release('ctrl + enter')
    time.sleep(3)

def send2PlayersHitters(player):

    keyboard.press_and_release('c')

    time.sleep(2)
    keyboard.write(player)
    time.sleep(1)

    keyboard.press_and_release('enter')

    time.sleep(1)

    keyboard.press_and_release('tab')

    time.sleep(1)
    keyboard.write('Postgame Report')

    time.sleep(1)
    keyboard.press_and_release('tab')
    keyboard.write('Here is tonights report.\n\n\nRyley')

    time.sleep(1)

    keyboard.press_and_release('tab')
    keyboard.press_and_release('tab')
    keyboard.press_and_release('tab')
    keyboard.press_and_release('enter')
    time.sleep(1)

    keyboard.write(f'{player} - Hitting')
    time.sleep(.25)

    time.sleep(1)
    keyboard.press_and_release('enter')
    time.sleep(3)
    keyboard.press_and_release('ctrl + enter')
    time.sleep(3)

    

def send2PlayersPitchers(player):

    keyboard.press_and_release('c')

    time.sleep(2)
    keyboard.write(player)
    time.sleep(1)

    keyboard.press_and_release('enter')

    time.sleep(1)

    keyboard.press_and_release('tab')

    time.sleep(1)
    keyboard.write('Postgame Report')

    time.sleep(1)
    keyboard.press_and_release('tab')
    keyboard.write('Here is tonights report.\n\n\nRyley')

    time.sleep(1)

    keyboard.press_and_release('tab')
    keyboard.press_and_release('tab')
    keyboard.press_and_release('tab')
    keyboard.press_and_release('enter')
    time.sleep(1)

    keyboard.write(f'{player} - Pitching')
    time.sleep(.25)

    time.sleep(1)
    keyboard.press_and_release('enter')
    time.sleep(3)
    keyboard.press_and_release('ctrl + enter')
    time.sleep(3)

def saveWorkbookHitter(wb, gamedate, player):
    
    hitterNames = player.split(',')
    hitter = hitterNames[1].strip() + ' ' + hitterNames[0]

    months = {1 : 'Jan', 2: 'Feb' , 3: 'Mar' , 4: 'Apr' , 5 : 'May' , 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12:'Dec'}
    if gamedate[4] == '-':
        dateSplit = gamedate.split('-')
        date = f'{months[int(dateSplit[1])]} {dateSplit[2]}'
    else:
        dateSplit = gamedate.split('/')
        date = f'{months[int(dateSplit[0])]} {dateSplit[1]}'
    
    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}')
        print(f'Adding folder: {date[:3]}')
    except:
        pass

    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}')
        try:
            os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Hitters')
        except:
            pass
    except:
        pass

    newPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Hitters\{hitter}-{date}.xlsx'
    wb.save(newPath)
    return newPath

def openWorkbook(gamedate, hitter):

    months = {1 : 'Jan', 2: 'Feb' , 3: 'Mar' , 4: 'Apr' , 5 : 'May' , 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12:'Dec'}
    if gamedate[4] == '-':
        dateSplit = gamedate.split('-')
        date = f'{months[int(dateSplit[1])]} {dateSplit[2]}'
    else:
        dateSplit = gamedate.split('/')
        date = f'{months[int(dateSplit[0])]} {dateSplit[1]}'

    path = f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Hitters\{hitter}-{date}.xlsx'

    wb = load_workbook(path)

    return wb.active, wb, path

def savePitcherWorkbook(wb, gamedate, pitcher):
    
    pitcherNames = pitcher.split(',')
    pitcher = pitcherNames[1].strip() + ' ' + pitcherNames[0]

    months = {1 : 'Jan', 2: 'Feb' , 3: 'Mar' , 4: 'Apr' , 5 : 'May' , 6: 'Jun', 7: 'Jul', 8: 'Aug', 9: 'Sep', 10: 'Oct', 11: 'Nov', 12:'Dec'}
    
    if gamedate[4] == '-':
        dateSplit = gamedate.split('-')
        date = f'{months[int(dateSplit[1])]} {dateSplit[2]}'
    else:
        dateSplit = gamedate.split('/')
        date = f'{months[int(dateSplit[0])]} {dateSplit[1]}'

    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}')
        print(f'Adding folder: {date[:3]}')
    except:
        pass
    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}')
    except:
        pass
    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Pitchers')
    except:
        pass

    newPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\\2022 Postgame Excel Sheets\{date[:3]}\{date}\Pitchers\{pitcher}-{date}.xlsx'
    wb.save(newPath)
    return newPath

def writePitcherReport(dbName, pitcher, date, tableName, opponent, autoTagPitchType):
    sheet, wb = openPitcherWorkbook()

    try:
        innings = getInningsPitched(dbName, pitcher, date, tableName, autoTagPitchType)
        writeOverall(sheet, dbName, pitcher, tableName, date, autoTagPitchType)
        writeHeaderPitcher(sheet, pitcher, date, opponent)
        imgCount = 13
        inningCount = 1
        altCount = True
        for inning in innings:
            writeInnings(sheet, dbName, pitcher, inning, inningCount, tableName, date, autoTagPitchType)
            if altCount:
                altCount = False
            else:
                altCount = True
            if imgCount == 68:
                imgCount += 1
            addStrikezoneFigure(dbName, pitcher, date, inning, imgCount, sheet, tableName, autoTagPitchType)
            inningCount += 1
            imgCount += 11
    except ZeroDivisionError:
        new = input('Please enter the pitcher exactly as it is entered in Trackman (or type "skip"): ')
        if new == 'skip':
            return
        innings = getInningsPitched(dbName, new, date, tableName, autoTagPitchType)
        writeOverall(sheet, dbName, new, tableName, date, autoTagPitchType)
        writeHeaderPitcher(sheet, new, date, opponent)
        imgCount = 13
        inningCount = 1
        altCount = True
        for inning in innings:
            writeInnings(sheet, dbName, new, inning, inningCount, tableName, date, autoTagPitchType)
            if altCount:
                altCount = False
            else:
                altCount = True
            if imgCount == 68:
                imgCount += 1
            addStrikezoneFigure(dbName, new, date, inning, imgCount, sheet, tableName, autoTagPitchType)
            inningCount += 1
            imgCount += 11

    if sheet['F78'].internal_value:
        sheet.print_area = 'B3:O132'
    else:
        sheet.print_area = 'B3:O76'

   
    path = savePitcherWorkbook(wb, gamedate, pitcher)
    printWorkbookPitcher(path, pitcher)
    
    return f'{pitcher.split()[1].rstrip(",")} {pitcher.split()[0].rstrip(",")}'

def writePlayerReport(dbName, player, gamedate, opponent, tableName, autoTagPitchType):

    sheet, wb = openHitterWorkbook()
    cur, con = openDatabase(dbName)
    saveWorkbookHitter(wb, gamedate, player)

    headerVars = 'pitchNum,hitter,date'
    writeHeader(headerVars, player, sheet, cur, opponent, tableName)

    pitchVars = f'pitchNum,{pitchType(autoTagPitchType)},pitchResult,locationX'
    writePitches(pitchVars, player, gamedate, sheet, cur, tableName)

    abVars = 'pitchNum,pitcher,outs,exitVelo,launchAngle,autoHitType,outcome,inning,ab,pitcherSide,topBottom,hitter,pitchResult,KorBB'
    writeAB(abVars, player, gamedate, sheet, cur, dbName, tableName, autoTagPitchType)

    if sheet['F57'].internal_value:
        sheet.print_area = 'C3:M96'
    else:
        sheet.print_area = 'C3:M49'

    #writeToQABSheet(f'{player.split()[1].rstrip(",")} {player.split()[0].rstrip(",")}', sheet, gamedate)

    path = saveWorkbookHitter(wb, gamedate, player)
    printWorkbookHitter(path, player)

    return f'{player.split()[1].rstrip(",")} {player.split()[0].rstrip(",")}'

def writeToQABSheet(player, report, date):

    path = "c:\\Users\\rlars\Desktop\Knights Baseball\Weekly QAB Report\Knights Weekly QAB Data.xlsx"
    wb = load_workbook(path)
    sheet = wb.active

    playerID = {'Jonah Advincula' : 1, 'Temo Becerra' : 2, 'Kyle Casperson' : 3, 'Zander Darby' : 4, 'Luca DiPaolo' : 5, 
    'Devon Dixon' : 6, 'Titus Dumitru' : 7, 'Logan Johnstone' : 8, 'Ely Kennel' : 9, 'Briley Knight' : 10,
    'Brady Lavoie' : 11, 'Mason Le' : 12, 'Ethan Loveless' : 13, 'Tyler Quinn' : 14, 'Kiko Romero' : 15, 'Spencer Scott' : 16, 'Tanner Smith' : 17, 'Ryan Stafford' : 18}

    dateID = {'2022-06-03':'C', '2022-06-04':'D','2022-06-05':'E',
    '2022-06-14':'F','2022-06-15':'G','2022-06-16':'H','2022-06-17':'I','2022-06-18':'J','2022-06-19':'K',
    '2022-06-21':'L','2022-06-22':'M','2022-06-23':'N','2022-06-24':'O','2022-06-25':'P','2022-06-26':'Q',
    '2022-06-28':'R','2022-06-29':'S','2022-06-30':'T','2022-07-01':'U','2022-07-02':'V','2022-07-03':'W',
    '2022-07-04':'X','2022-07-05':'Y','2022-07-06':'Z','2022-07-07':'AA','2022-07-08':'AB','2022-07-09':'AC','2022-07-10':'AD',
    '2022-07-13':'AE','2022-07-15':'AF','2022-07-16':'AG','2022-07-17':'AH',
    '2022-07-18':'AI','2022-07-19':'AJ','2022-07-20':'AK','2022-07-21':'AL','2022-07-22':'AM', '2022-07-23' : 'AN', '2022-07-24' : 'AO',
    '2022-07-25' : 'AP', '2022-07-26' : 'AQ', '2022-07-27' : 'AR', '2022-07-28' : 'AS', '2022-07-29' : 'AT', '2022-07-30' : 'AU', '2022-07-31' : 'AV'}

    buffer = 100
    header = 5
    overall = 50

    qabList = []
    throwsList = []
    abCount = 1
    count = 16
    
    while report[f'L{count}'].internal_value:
        qabList.append(report[f'L{count}'].internal_value)
        throwsList.append(report[f'I{count-2}'].internal_value)
        count += 8
        abCount += 1
        if abCount > 5:
            count += 7
    
    yesR, totR = 0, 0
    yesL, totL = 0, 0
    yesO, totO = 0, 0
    for i in range(len(qabList)):
        if throwsList[i] == 'Right':
            if qabList[i] == 'Yes':
                yesR += 1
            totR += 1
        elif throwsList[i] == 'Left':
            if qabList[i] == 'Yes':
                yesL += 1
            totL += 1
        elif throwsList[i] == 'Undefined':
            if qabList[i] == 'Yes':
                yesO += 1
            totO += 1
        else:
            print('QAB, but no Throws')


    num = 2*playerID[player] + header

    sheet[f"{dateID[date]}{num}"] = (yesR + yesL + yesO)
    sheet[f"{dateID[date]}{num+1}"] = (totR+totL+totO)

    num += overall

    sheet[f"{dateID[date]}{num}"] = yesL
    sheet[f"{dateID[date]}{num+1}"] = totL

    num += buffer

    sheet[f"{dateID[date]}{num}"] = yesR
    sheet[f"{dateID[date]}{num+1}"] = totR

    wb.save(path)

def iterateOverGamesQABs(lineup):

    games = [('2022-06-03', 'YakimaValley'), ('2022-06-04', 'YakimaValley'),('2022-06-05', 'YakimaValley'),
    ('2022-06-14', 'Cowlitz'),('2022-06-15', 'Cowlitz'),('2022-06-16', 'Cowlitz'),('2022-06-17', 'YakimaValley'),('2022-06-18', 'YakimaValley'),('2022-06-19', 'YakimaValley'),
    ('2022-06-21', 'WallaWalla'),('2022-06-22', 'WallaWalla'),('2022-06-23', 'WallaWalla'),('2022-06-24', 'Bellingham'),('2022-06-25', 'Bellingham'),('2022-06-26', 'Bellingham'),
    ('2022-06-28', 'Springfield'),('2022-06-29', 'Springfield'),('2022-06-30', 'Springfield'),('2022-07-01', 'PortAngeles'),('2022-07-02', 'PortAngeles'),('2022-07-03', 'PortAngeles'),
    ('2022-07-04', 'Portland'),('2022-07-05', 'Ridgefield'),('2022-07-06', 'Ridgefield'),('2022-07-07', 'Ridgefield'),('2022-07-08', 'Bend'),('2022-07-09', 'Bend'),('2022-07-10', 'Bend'),
    ('2022-07-13', 'Edmonton'),('2022-07-15', 'Wenatchee'),('2022-07-16', 'Wenatchee'),('2022-07-17', 'Wenatchee'),
    ('2022-07-18', 'Portland'),('2022-07-19', 'Cowlitz'),('2022-07-20', 'Cowlitz'),('2022-07-21', 'Cowlitz'),('2022-07-22', 'Portland')]


    for game in games:
        gamedate = game[0]

        for player in lineup:
            player = f'{player.split()[1].rstrip(",")} {player.split()[0].rstrip(",")}'
            report, wb, path = openWorkbook(gamedate, player)
            writeToQABSheet(player, report, gamedate)
            wb.save(path)

def main(lineup, dbName, gamedate, opponent, tableName, pitchers, autoTagPitchType, path, aX = adjustmentX, aY = adjustmentY):
    try:
        writeDatabase(dbName, tableName, path, aX, aY)
        print('...writing database...')
    except sqlite3.IntegrityError:
        print('...database already written...')
    if autoTagPitchType:
        print('...autoPitch on...')
    else:
        print('...autoPitch off...')
    players = []
    for player in lineup:
        try:
            hitter = writePlayerReport(dbName, player, gamedate, opponent, tableName, autoTagPitchType)
            players.append(hitter)
        except IndexError:
            #print(f'No report written for {player}')
            continue
    plt.close('all')
    throwers = []
    for pitcher in pitchers: 
        try:
            thrower = writePitcherReport(dbName, pitcher, gamedate, tableName, opponent, autoTagPitchType)
            throwers.append(thrower)
        except IndexError:
            print(f'No report written for {pitcher}')
            continue
    plt.close('all')

    keyboard.press_and_release('Windows + m')
    time.sleep(.5)
    keyboard.press_and_release('ctrl + alt + g')
    time.sleep(5.5)

    for player in players:
        time.sleep(2)
        send2PlayersHitters(player)
        time.sleep(2)
    for thrower in throwers:
        time.sleep(2)
        send2PlayersPitchers(thrower)
        time.sleep(2)
    sendHitterWorkbooks(players)
    sendPitcherWorkbooks(throwers)
    keyboard.press_and_release('Windows + m')

def send2Coaches(players, pitchers):
    
    keyboard.press_and_release('Windows + m')
    time.sleep(2)
    keyboard.press_and_release('ctrl + alt + g')
    time.sleep(6)
    keyboard.press_and_release('fn + f11')

    sendPitcherWorkbooks(pitchers)
    sendHitterWorkbooks(players)
    
    keyboard.press_and_release('Windows + m')
    
main(lineup, dbName, gamedate, opponent, tableName, pitchers, autoTagPitchType, path, aX = adjustmentX, aY = adjustmentY)

