import os
import sqlite3
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font, PatternFill
import matplotlib.pyplot as plt
from matplotlib.patches import Rectangle
import pandas as pd

def openDatabase():
    path = "c:\\Users\\rlars\Desktop\Knights Baseball\Postgame Reports\Databases\Knights2022"
    con = sqlite3.connect(path)
    cur = con.cursor()

    return cur, con

def openWorkbook():
    copyPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Hitter Chase Report\Hitter Chase Template.xlsx'
    wb = load_workbook(copyPath)
    return wb.active, wb

def getChase(pitch):
    chase = False
    ball = False
    if pitch[0]:
        if pitch[0] < -0.708333333333333 or pitch[0] > 0.7083333333333333:
            ball = True
        if pitch[1] < 1.55 or pitch[1] > (1.55+1.67):
            ball = True
        if ball and pitch[3] == ('StrikeSwinging' or 'InPlay' or 'FoulBall'):
            chase = True
    return chase

def getChaseCount(pitch, chaseCount):
    ball = False
    if pitch[0]:
        if pitch[0] < -0.708333333333333 or pitch[0] > 0.7083333333333333:
            ball = True
        if pitch[1] < 1.55 or pitch[1] > (1.55+1.67):
            ball = True
        if ball and pitch[3] == ('StrikeSwinging' or 'InPlay' or 'FoulBall'):
            chaseCount += 1
    return chaseCount

def getBall(pitch):
    ball = False
    if pitch[0]:
        if pitch[0] < -0.708333333333333 or pitch[0] > 0.7083333333333333:
            ball = True
        if pitch[1] < 1.55 or pitch[1] > (1.55+1.67):
            ball = True
    return ball

def getTypePitch(typePitch, pitch):
    fastballs = ['Fastball', 'Four-Seam', 'Sinker']

    if typePitch == 'Fastball':
        if pitch[2] in fastballs:
            return True
        else:
            return False
    else:
        if pitch[2] in fastballs:
            return False
        else:
            return True

def getAllPitches(player):
    cur, con = openDatabase()
    tableNames = ['June17vYakimaValley', 'June18vYakimaValley', 'June19vYakimaValley', 'June21vWallaWalla', 'June22vWallaWalla', 'June23vWallaWalla', 'June24vBellingham', 'June25vBellingham', 'June26vBellingham', 'June28vSpringfield', 'June29vSpringfield', 'June30vSpringfield', 'July01vPortAngeles', 'July02vPortAngeles']
    ret = []
    for tableName in tableNames:
        try:
            cur.execute(f'SELECT locationX, locationY, taggedPitch, pitchResult, strikes FROM {tableName} WHERE hitter = ?' , (player,))
            allPitches = cur.fetchall()
            ret.extend(allPitches)
        except:
            pass

    return ret

def getChaseRate(typePitch, allPitches):

    typePitches = [tup for tup in allPitches if getTypePitch(typePitch, tup)]
    earlyPitches = [tup for tup in typePitches if tup[4] != 2]

    chaseCount = 0
    pitchCount = 0
    for pitch in earlyPitches:
        chaseCount = getChaseCount(pitch, chaseCount)
        if getBall(pitch):
            pitchCount += 1
    return (chaseCount / pitchCount)

def get2KChaseRate(typePitch, allPitches):

    typePitches = [tup for tup in allPitches if getTypePitch(typePitch, tup)]
    kPitches = [tup for tup in typePitches if tup[4] == 2]

    chaseCount = 0
    pitchCount = 0
    for pitch in kPitches:
        chaseCount = getChaseCount(pitch, chaseCount)
        if getBall(pitch):
            pitchCount += 1
    return (chaseCount / pitchCount)

def drawStrikezoneFigure(player, typePitch, allPitches):

    s = .75

    fig, ax = plt.subplots(figsize = (3/s, 6/s))
    ax.set_xlim([(-1.5/s),(1.5/s)])
    ax.set_ylim([0/s,6/s])
    ax.spines['top'].set_visible(False)
    ax.spines['right'].set_visible(False)
    ax.spines['bottom'].set_visible(False)
    ax.spines['left'].set_visible(False)
    ax.axes.xaxis.set_visible(False)
    ax.axes.yaxis.set_visible(False)

    typePitches = [tup for tup in allPitches if getTypePitch(typePitch, tup)]
    chasePitches = [tup for tup in typePitches if getChase(tup)]

    x = []
    y = []
    n = []
    c = []

    num = 1
    for pitch in chasePitches:
        if not pitch[0]:
            continue
        if ((pitch[0]*-1)/s) >= 1.4/s:
            x.append(1.4/s)
        elif ((pitch[0]*-1)/s) <= -1.4/s:
            x.append(-1.4/s)
        else:
            x.append((pitch[0]*-1)/s)
        if (pitch[1]/s) >= 3.9/s:
            y.append(3.9/s)
        elif (pitch[1]/s) <= 1.25/s:
            y.append(1.25/s)
        else:
            y.append(pitch[1]/s)
        n.append(num)
        c.append(pitch[4])
        num += 1

    ax.add_patch(Rectangle((-0.708333333333/s, 1.55/s), 1.416666666667/s, 1.67/s,
             edgecolor = 'black', facecolor= '#98012E',
             fill=True,
             lw = 1)),

    df = pd.DataFrame({'x': x,
                   'y': y,
                   'c': c})

    strikesDic = {0 : 'Black', 1 : 'Black', 2 : '#98012E'}
    groups = df.groupby('c')
    for name, group in groups:
        plt.scatter(group.x, group.y, s = 100/s, zorder = 3, color = strikesDic[name], edgecolors= 'white', linewidths = .45)

    try:
        os.mkdir(f'c:\\Users\\rlars\Desktop\Knights Baseball\Hitter Chase Report\Graphics')
    except:
        pass
    newPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Hitter Chase Report\Graphics\{player}-{typePitch}.png'

    
    plt.savefig(newPath, transparent = True)
    plt.close()

    return Image(newPath)

def addStrikeZone(player, typePitch, sheet, allPitches):
    anchorDic = {'Fastball' : 'E1', 'Offspeed' : 'L1'}

    img = drawStrikezoneFigure(player, typePitch, allPitches)
    sheet.add_image(img, anchorDic[typePitch])

def saveWorkbook(wb, player):
    
    hitterNames = player.split(',')
    hitter = hitterNames[1].strip() + ' ' + hitterNames[0]

    newPath = f'c:\\Users\\rlars\Desktop\Knights Baseball\Hitter Chase Report\Excel Sheets\{hitter}.xlsx'
    wb.save(newPath)
    return newPath

def writeWorkbook(player, allPitches):
    sheet, wb = openWorkbook()

    sheet[f'D8'] = player.split(',')[0]

    pitchTypes = {'Fastball': 'H33', 'Offspeed': 'O33'}
    pitchTypes2k = {'Fastball': 'H36', 'Offspeed': 'O36'}

    for typePitch in pitchTypes:
        sheet[pitchTypes[typePitch]] = getChaseRate(typePitch, allPitches)
        sheet[pitchTypes2k[typePitch]] = get2KChaseRate(typePitch, allPitches)
        addStrikeZone(player, typePitch, sheet, allPitches)

    return saveWorkbook(wb, player)

def printWorkbook(path, player):
    if os.path.isfile(path):
        try:
            i = input(f'Ready to print report: {player} \nPress enter to continue: ')
            if i == ' ':
                return None
            else:
                os.startfile(path, 'print')
        except:
            print(f'Exception occured: {player}')
    else:
        print(f'Error with file: {player}')

def main():

    players = ['Johnstone, Logan', 'Scott, Spencer', 'Advincula, Jonah', 'Kennel, Ely', 
    'Le, Mason', 'Darby, Zander', 'Lavoie, Brady', 'Dumitru, Titus', 'Quinn, Tyler', 'Romero, Kiko', 
    'Loveless, Ethan', 'Casperson, Kyle', 'Leitgeb, Jake', 'Becerra, Temo', 'Knight, Briley', 'DiPaolo, Luca']

    for player in players:
        allPitches = getAllPitches(player)
        path = writeWorkbook(player, allPitches)
        printWorkbook(path, player)

main()